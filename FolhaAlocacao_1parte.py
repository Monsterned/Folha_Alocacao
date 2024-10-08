import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
import os
import tkinter as tk
from tkinter import messagebox

# Carregar o Excel especificando que o cabeçalho está na segunda linha (índice 1)
Base = pd.read_excel("Romaneio de Descarga carga 202191.xlsx", header=2)
Base = Base.dropna(subset=['CNPJ'])
Base['Cod. Produto'] = Base['Cod. Produto'].astype(int)
ordem = Base['Ordem'].iloc[0]
ordem = int(ordem)

# Agrupar pelo 'Cod. Produto' e somar a coluna 'Qtde', mantendo a primeira descrição encontrada
df_agrupado = Base.groupby('Cod. Produto').agg({'Descrição': 'first', 'Qtde': 'sum'}).reset_index()

# Adicionar novas colunas ao DataFrame agrupado
df_agrupado['END1'] = ''
df_agrupado['END2'] = ''
df_agrupado['END3'] = ''
df_agrupado['END4'] = ''
df_agrupado['END5'] = ''
df_agrupado['OBS'] = ''

# Calcular a soma da coluna 'Qtde'
soma_qtde = df_agrupado['Qtde'].sum()

# Criar uma linha de total com o mesmo formato
total_row = pd.DataFrame({
    'Cod. Produto': ['Total Geral'],
    'Descrição': [''],
    'Qtde': [soma_qtde],
    'END1': [''],
    'END2': [''],
    'END3': [''],
    'END4': [''],
    'END5': [''],
    'OBS': ['']
})

# Adicionar a linha de total ao DataFrame
df_agrupado = pd.concat([df_agrupado, total_row], ignore_index=True)

# Adicionar as novas colunas "ALOCAÇÃO", "ALOCAÇÃO", e "ROTA" ao DataFrame original "Base"
Base['ALOCAÇÃO1'] = ''
Base['ALOCAÇÃO2'] = ''
Base['ROTA'] = ''

# Exportar para um arquivo Excel
file_path = 'CACAU SHOW 202191 RESUMIDA.xlsx'
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    # Salvar a planilha original com novas colunas na aba "BASE"
    Base.to_excel(writer, sheet_name='BASE', index=False)
    
    # Salvar a planilha modificada na aba "ALOCAÇÃO"
    df_agrupado.to_excel(writer, sheet_name='ALOCAÇÃO', index=False)

# Carregar o arquivo Excel e a planilha "ALOCAÇÃO"
wb = load_workbook(file_path)
ws = wb['ALOCAÇÃO']

# Configurar a orientação da página para paisagem
ws.page_setup.orientation = 'landscape'

# Ajustar margens da página
ws.page_margins = PageMargins(left=0.75, right=0.75, top=0.75, bottom=0.75, header=0.3, footer=0.3)

# Definir a escala de ajuste para caber em uma página
ws.page_setup.fitToPage = True
ws.page_setup.fitToHeight = 0  # Ajuste para caber em uma página de altura
ws.page_setup.fitToWidth = 1   # Ajuste para caber em uma página de largura

# Adicionar texto ao cabeçalho e rodapé
ws.oddHeader.center.text = "Página &P de &N"  # &P é o número da página e &N é o número total de páginas
ws.oddFooter.center.text = "Gerado por Meu Script"  # Exemplo de texto adicional no rodapé

# Ajustar a largura das colunas e adicionar bordas
for col_num, value in enumerate(df_agrupado.columns.values):
    column_length = df_agrupado[value].astype(str).map(len).max()
    column_length = max(column_length, len(value)) + 2  # +2 para espaçamento extra
    col_letter = chr(65 + col_num)  # Convertendo número da coluna para letra (A, B, C, ...)
    ws.column_dimensions[col_letter].width = column_length

# Ajustar a largura das colunas específicas
column_widths = {
    'D': 10,  # Largura desejada para a coluna D
    'E': 10,  # Largura desejada para a coluna E
    'F': 10,  # Largura desejada para a coluna F
    'G': 10,  # Largura desejada para a coluna G
    'H': 10,  # Largura desejada para a coluna H
    'I': 10   # Largura desejada para a coluna I
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Adicionar bordas a todas as células
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        cell.border = border

# Ajustar a altura das linhas
row_height = 26  # Defina a altura desejada para as linhas
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = row_height

# Centralizar o texto nas colunas A, B e C
for col in ['A', 'B', 'C']:
    for cell in ws[col]:
        cell.alignment = Alignment(horizontal='center')

# Centralizar o texto na célula "Total"
total_cell = ws[f'A{len(df_agrupado)}']
total_cell.alignment = Alignment(horizontal='center')

# Garantir que as linhas de grade estejam visíveis (para configuração do Excel)
ws.sheet_view.show_grid_lines = True

# Adicionar nova informação três linhas abaixo da última linha
nova_informacao1 = f'CACAU SHOW {ordem}'  # Substitua pelo valor que deseja adicionar
linha_nova = len(df_agrupado) + 3  # Calcula a linha três linhas abaixo da última
ws[f'A{linha_nova}'] = nova_informacao1

# Adicionar nova informação na coluna 'C' três linhas abaixo da última linha
nova_informacao2 = 'ALOCAÇÃO'  # Substitua pelo valor que deseja adicionar
ws[f'C{linha_nova}'] = nova_informacao2

# Adicionar mais uma nova informação na coluna 'G' três linhas abaixo da última linha
nova_informacao3 = '26/08/2024'  # Substitua pelo valor que deseja adicionar
ws[f'G{linha_nova}'] = nova_informacao3

# Salvar as alterações
wb.save(file_path)

print("Arquivo Excel criado com a aba BASE contendo novas colunas e a aba ALOCAÇÃO modificada.")

# Função que será chamada para exibir a mensagem
def show_success_message():
    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal
    messagebox.showinfo("Sucesso", "Código finalizado com sucesso!")
    root.destroy()  # Fecha a janela após exibir a mensagem

# Simulando o final do código
if __name__ == "__main__":
    # Aqui você pode adicionar seu código que será executado
    # e ao final, chamará a função para exibir a mensagem
    show_success_message()