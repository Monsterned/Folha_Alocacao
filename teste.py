from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins

# Carregar o arquivo Excel existente
arquivo = 'CACAU SHOW 202191 RESUMIDA.xlsx'
wb = load_workbook(arquivo)
ws = wb['DINAMICA']  # Nome da aba onde está a coluna ALOCAÇÃO

# Ajustar a largura da coluna ALOCAÇÃO
ws.column_dimensions['E'].width = 37  # Ajuste a largura conforme necessário

# Ajustar a altura da linha e quebrar o texto na coluna ALOCAÇÃO
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.column_letter == 'E':  # Substitua 'E' pela letra da coluna ALOCAÇÃO, se necessário
            cell.alignment = Alignment(wrap_text=True)  # Garantir que wrap_text seja aplicado
            ws.row_dimensions[cell.row].height = 70  # Ajustar a altura da linha (mude conforme necessário)

# Configurar para ajustar o conteúdo à largura da página A4
ws.page_setup.fitToWidth = 1  # Ajustar para caber na largura de uma página
ws.page_setup.fitToHeight = 0  # Manter altura automática, pode ocupar várias páginas verticalmente
ws.page_setup.orientation = 'landscape'  # Opcional: mudar para paisagem se necessário

# Margens da página (opcional)
ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75)

# Salvar o arquivo Excel
wb.save(arquivo)