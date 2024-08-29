import pandas as pd
import openpyxl
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Border, Side
import numpy as np
import os
from openpyxl.styles import Alignment
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins

caminho = os.getcwd() 
data_hoje = datetime.now().strftime('%d/%m/%Y')

# Carregar os DataFrames
BASE = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='BASE')
ALOCAÇÃO = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='ALOCAÇÃO')
ROTAS = pd.read_excel('ROTAS TERMINAL CACAU.xlsx', sheet_name='Plan1', usecols="A:B")

ordem = BASE['Ordem'].iloc[0]
ordem = int(ordem)

# Realizar o merge para buscar a coluna desejada em ALOCAÇÃO
resultado = pd.merge(BASE, ALOCAÇÃO.iloc[:, [0, 3]], how='left', left_on='Cod. Produto', right_on=ALOCAÇÃO.columns[0])
BASE['ALOCAÇÃO1'] = resultado.iloc[:, -1]

# Criar a nova coluna 'ALOCAÇÃO2' com os três primeiros caracteres da coluna 'ALOCAÇÃO1'
BASE['ALOCAÇÃO2'] = BASE['ALOCAÇÃO1'].str[:3]

# Ajustar o nome da coluna de pesquisa conforme necessário
BASE = pd.merge(BASE, ROTAS, how='left', left_on='ALOCAÇÃO2', right_on=ROTAS.columns[0])

# Renomear a coluna resultante para 'ROTA' (ou o nome desejado)
BASE.rename(columns={ROTAS.columns[1]: 'ROTA'}, inplace=True)

# Remover a coluna de pesquisa se não for mais necessária
BASE.drop(columns=[ROTAS.columns[0]], inplace=True)

# Salvar o DataFrame atualizado de volta em um arquivo Excel na guia BASE
with pd.ExcelWriter('CACAU SHOW 202191 RESUMIDA.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    BASE.to_excel(writer, sheet_name='BASE', index=False)

# Carregar os DataFrames
BASE = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='BASE')
ALOCAÇÃO = pd.read_excel('CACAU SHOW 202191 RESUMIDA.xlsx', sheet_name='ALOCAÇÃO')
ROTAS = pd.read_excel('ROTAS TERMINAL CACAU.xlsx', sheet_name='Plan1', usecols="A:B")

# Realizar o merge para buscar a coluna desejada em ALOCAÇÃO
resultado = pd.merge(BASE, ALOCAÇÃO.iloc[:, [0, 3]], how='left', left_on='Cod. Produto', right_on=ALOCAÇÃO.columns[0])
BASE['ALOCAÇÃO1'] = resultado.iloc[:, -1]

# Criar a nova coluna 'ALOCAÇÃO2' com os três primeiros caracteres da coluna 'ALOCAÇÃO1'
BASE['ALOCAÇÃO2'] = BASE['ALOCAÇÃO1'].str[:3]

# Realizar o merge com o DataFrame ROTAS para adicionar a coluna desejada
BASE = pd.merge(BASE, ROTAS, how='left', left_on='ALOCAÇÃO1', right_on=ROTAS.columns[0])

# Renomear a coluna resultante para 'ROTA'
BASE.rename(columns={ROTAS.columns[1]: 'ROTA'}, inplace=True)

# Remover a coluna de pesquisa se não for mais necessária
BASE.drop(columns=[ROTAS.columns[0]], inplace=True)

# Criar a Tabela Dinâmica
tabela_dinamica = BASE.pivot_table(
    index=['Cte', 'Cidade', 'ALOCAÇÃO1', 'Descrição', 'Cod. Produto'],
    values='Qtde',
    aggfunc='sum',
    fill_value=0
).reset_index()

# Ajustar o tipo da coluna 'Cte' para string
tabela_dinamica['Cte'] = tabela_dinamica['Cte'].astype(str)

# Adicionar totais antes de cada novo Cte
def add_totals(df):
    result = []
    for cte in df['Cte'].unique():
        subset = df[df['Cte'] == cte]
        total_row = subset[['Qtde']].sum()
        total_row = total_row.to_frame().T
        total_row['Cte'] = ''
        total_row['Cidade'] = ''
        total_row['Descrição'] = ''
        total_row['Cod. Produto'] = ''
        total_row['ALOCAÇÃO1'] = ''
        result.append(subset)
        result.append(total_row)
    return pd.concat(result, ignore_index=True)

# Adicionar totais
tabela_dinamica = add_totals(tabela_dinamica)

# Salvar o DataFrame atualizado e a Tabela Dinâmica de volta em um arquivo Excel
with pd.ExcelWriter('CACAU SHOW 202191 RESUMIDA.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    BASE.to_excel(writer, sheet_name='BASE', index=False)
    tabela_dinamica.to_excel(writer, sheet_name='DINAMICA', index=False)

arquivo_xlsx = caminho + r'\CACAU SHOW 202191 RESUMIDA.xlsx'

# Carregar a planilha convertida
Planilha_xml = pd.read_excel(arquivo_xlsx, sheet_name="DINAMICA")
Planilha_xml = Planilha_xml.rename(columns={'ALOCAÇÃO1': 'ALOCAÇÃO'})
# Substituir valores vazios na coluna 'Cod. Produto' por NaN
Planilha_xml['Cod. Produto'] = Planilha_xml['Cod. Produto'].replace('', np.nan)

# Substituir NaN por strings vazias
Planilha_xml = Planilha_xml.fillna('')

# Salvar o DataFrame modificado de volta para o arquivo Excel
with pd.ExcelWriter(arquivo_xlsx, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    Planilha_xml.to_excel(writer, sheet_name="DINAMICA", index=False)

print("Valores vazios na coluna 'Cod. Produto' substituídos por células em branco!")









# Adicionar a quebra de página e limpar células
wb = openpyxl.load_workbook(arquivo_xlsx)
ws = wb["DINAMICA"]

# Adicionar quebras de página e limpar células onde 'Cod. Produto' está vazio
for idx, row in Planilha_xml.iterrows():
    if row['Cod. Produto'] == '':
        # Limpar células nas colunas 'Cte' e 'Soma de Qtde'
        ws[f'A{idx + 2}'].value = None
        # Adicionar quebra de página após a linha onde 'Cod. Produto' está vazio
        ws.row_breaks.append(Break(id=idx + 2))  # +2 para adicionar a quebra após a linha com 'Cod. Produto' vazio

# Ajustar a largura das colunas específicas
colunas_para_ajustar = ['Cidade', 'Cod. Produto','Cod. Produto', 'Qtde']

for coluna in colunas_para_ajustar:
    col_index = Planilha_xml.columns.get_loc(coluna) + 1  # +1 porque openpyxl é 1-based index
    col_letter = openpyxl.utils.get_column_letter(col_index)

    max_length = 0
    for cell in ws[col_letter]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass

    # Definir a largura da coluna
    ws.column_dimensions[col_letter].width = max_length + 2  # +2 para algum espaçamento extra

# Ajustar a largura da coluna 'Descrição'
descricao_col = 'Descrição'
descricao_index = Planilha_xml.columns.get_loc(descricao_col) + 1
descricao_letter = openpyxl.utils.get_column_letter(descricao_index)

descricao_max_length = 0
for cell in ws[descricao_letter]:
    try:
        if len(str(cell.value)) > descricao_max_length:
            descricao_max_length = len(cell.value)
    except:
        pass

# Definir a largura da coluna 'Descrição'
ws.column_dimensions[descricao_letter].width = descricao_max_length + 2  # +2 para algum espaçamento extra

# Adicionar bordas horizontais às linhas
border_top = Side(style='thin')
border_bottom = Side(style='thin')
border_horizontal = Border(top=border_top, bottom=border_bottom)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Começa a partir da segunda linha para evitar o cabeçalho
    for cell in row:
        cell.border = border_horizontal

# Configurar a orientação da página para paisagem
ws.page_setup.orientation = 'landscape'

# Ajustar a altura das linhas
row_height = 22  # Defina a altura desejada para as linhas
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = row_height

# Configurar cabeçalhos e rodapés para incluir numeração de páginas
ws.oddHeader.center.text = "Página &P de &N"
ws.evenHeader.center.text = "Página &P de &N"

# Adicionar o texto "CACAU SHOW" e a data na última linha
ultima_linha = ws.max_row + 1
ws[f'A{ultima_linha}'] = f"CACAU SHOW {ordem}"
ws[f'C{ultima_linha}'] = f"ALOCAÇÃO"
ws[f'E{ultima_linha}'] = data_hoje

# Salvar o arquivo Excel com as quebras de página, células limpas, larguras das colunas ajustadas, bordas horizontais e cabeçalhos/rodapés configurados
wb.save(arquivo_xlsx)


# Carregar o arquivo Excel existente
arquivo = 'CACAU SHOW 202191 RESUMIDA.xlsx'
wb = load_workbook(arquivo)
ws = wb['DINAMICA']  # Nome da aba onde está a coluna ALOCAÇÃO

# Ajustar a largura da coluna ALOCAÇÃO
ws.column_dimensions['E'].width = 37  # Ajuste a largura conforme necessário

# Ajustar a altura da linha e quebrar o texto na coluna ALOCAÇÃO
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.column_letter == 'E':  # Substitua 'E' pela letra da coluna ALOCAÇÃO, se necessário
            cell.alignment = Alignment(wrap_text=True)  # Garantir que wrap_text seja aplicado
            ws.row_dimensions[cell.row].height = 70  # Ajustar a altura da linha (mude conforme necessário)

# Configurar para ajustar o conteúdo à largura da página A4
ws.page_setup.fitToWidth = 1  # Ajustar para caber na largura de uma página
ws.page_setup.fitToHeight = 0  # Manter altura automática, pode ocupar várias páginas verticalmente
ws.page_setup.orientation = 'landscape'  # Opcional: mudar para paisagem se necessário

# Margens da página (opcional)
ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75)

# Salvar o arquivo Excel
wb.save(arquivo)



print("Quebras de página adicionadas, células limpas, larguras das colunas ajustadas, bordas horizontais adicionadas e cabeçalhos/rodapés configurados com numeração de páginas!")
print('Terminou :)')